import argparse
import sys
from pathlib import Path
from datetime import datetime, date
from openpyxl import load_workbook

def dart_escape(s: str) -> str:
    return s.replace("\\", "\\\\").replace("'", "\\'")

def dart_value(v) -> str:
    if v is None:
        return "null"
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        return str(v)
    if isinstance(v, (datetime, date)):
        return "'" + v.isoformat() + "'"
    return "'" + dart_escape(str(v)) + "'"

def read_rows(ws, header_row: int):
    headers = []
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=c).value
        headers.append(str(val).strip() if val is not None else "")
    data = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = {}
        empty = True
        for c, h in enumerate(headers, start=1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip() != "":
                empty = False
            if h != "":
                row[h] = v
        if not empty:
            data.append(row)
    return headers, data

def generate_dart(headers, rows, key_field: str, variable: str) -> str:
    keyed = {}
    for row in rows:
        key_raw = row.get(key_field)
        if key_raw is None:
            continue
        key = str(key_raw).strip()
        keyed[key] = row
    keys_sorted = sorted(keyed.keys(), key=lambda s: s)
    lines = []
    lines.append(f"const Map<String, Map<String, Object?>> {variable} = <String, Map<String, Object?>>{{")
    for key in keys_sorted:
        row = keyed[key]
        lines.append(f"  '{dart_escape(key)}': <String, Object?>{{")
        lines.append(f"    '{dart_escape(key_field)}': {dart_value(row.get(key_field))},")
        for h in headers:
            if h == "" or h == key_field:
                continue
            if h in row:
                lines.append(f"    '{dart_escape(h)}': {dart_value(row.get(h))},")
        extras = [k for k in row.keys() if k not in headers and k != key_field and k != ""]
        for k in sorted(extras):
            lines.append(f"    '{dart_escape(k)}': {dart_value(row.get(k))},")
        lines.append("  },")
    lines.append("};")
    return "\n".join(lines)

def main():
    ap = argparse.ArgumentParser(prog="excel_to_dart")
    ap.add_argument("-i", "--input", required=True)
    ap.add_argument("-o", "--output", required=True)
    ap.add_argument("-s", "--sheet", default=None)
    ap.add_argument("--header-row", type=int, default=1)
    ap.add_argument("--key", default="name")
    ap.add_argument("--variable", default="coursesByName")
    args = ap.parse_args()
    ip = Path(args.input)
    if not ip.exists():
        print("输入文件不存在", file=sys.stderr)
        sys.exit(1)
    wb = load_workbook(filename=str(ip), data_only=True)
    ws = wb[args.sheet] if args.sheet else wb.active
    headers, rows = read_rows(ws, args.header_row)
    if args.key not in headers:
        print("关键字段不存在", file=sys.stderr)
        sys.exit(1)
    content = generate_dart(headers, rows, args.key, args.variable)
    op = Path(args.output)
    op.parent.mkdir(parents=True, exist_ok=True)
    op.write_text(content, encoding="utf-8")
    print(str(op))

if __name__ == "__main__":
    main()
