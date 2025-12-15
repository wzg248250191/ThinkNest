import argparse
import sys
from pathlib import Path
from openpyxl import load_workbook
from pypinyin import lazy_pinyin, Style

def initials(text: str, mode: str = "initials") -> str:
    s = "".join(lazy_pinyin(str(text), style=Style.FIRST_LETTER)).lower()
    return s[:1] if mode == "first" else s

def last_used_column(ws, row: int) -> int:
    last = 0
    mc = ws.max_column
    for c in range(1, mc + 1):
        v = ws.cell(row=row, column=c).value
        if v is not None and str(v).strip() != "":
            last = c
    return last

def process(input_path: Path, output_path: Path | None, sheet: str | None, start_row: int, mode: str) -> Path:
    wb = load_workbook(filename=str(input_path))
    ws = wb[sheet] if sheet else wb.active
    max_row = ws.max_row
    for r in range(start_row, max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is None or str(v).strip() == "":
            continue
        val = initials(v, mode)
        if val == "":
            continue
        last = last_used_column(ws, r)
        target_col = last + 1 if last >= 1 else 1
        ws.cell(row=r, column=target_col, value=val)
    out = output_path if output_path else input_path
    wb.save(str(out))
    return out

def main():
    ap = argparse.ArgumentParser(prog="pinyin_excel", description="将第一列中文的首字母写入该行最后一列")
    ap.add_argument("-i", "--input", required=True, help="输入xlsx文件路径")
    ap.add_argument("-o", "--output", default=None, help="输出xlsx文件路径(默认覆盖输入文件)")
    ap.add_argument("-s", "--sheet", default=None, help="工作表名称(默认活动表)")
    ap.add_argument("--start-row", type=int, default=1, help="开始行号(默认1)")
    ap.add_argument("--mode", choices=["initials", "first"], default="initials", help="initials=全串首字母, first=仅首字母")
    args = ap.parse_args()
    ip = Path(args.input)
    if not ip.exists():
        print("输入文件不存在", file=sys.stderr)
        sys.exit(1)
    op = Path(args.output) if args.output else None
    try:
        out = process(ip, op, args.sheet, args.start_row, args.mode)
        print(str(out))
    except Exception as e:
        print(f"处理失败: {e}", file=sys.stderr)
        sys.exit(2)

if __name__ == "__main__":
    main()
